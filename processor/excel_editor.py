from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os
from PIL import Image as PILImage

def debug_excel_content(ws):
    """Debug específico para encontrar {{LOGO}}"""
    print(f"🔍 Buscando marcadores en Excel...")
    found_markers = False
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_text = str(cell.value)
                if "LOGO" in cell_text or "{{" in cell_text:
                    print(f"   📍 Encontrado en {cell.coordinate}: '{cell_text}'")
                    found_markers = True
    
    if not found_markers:
        print("   ⚠️ No se encontraron marcadores {{LOGO}}")
    
    return found_markers

def replace_text_in_excel(workbook, replacements: dict):
    """Reemplaza texto en todas las hojas del workbook."""
    print("📝 Procesando reemplazos de texto en Excel...")
    replaced_count = 0
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"   📄 Procesando hoja: {sheet_name}")
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_text = cell.value
                    modified_text = original_text
                    
                    # Aplicar reemplazos
                    for key, value in replacements.items():
                        if key in modified_text:
                            modified_text = modified_text.replace(key, value)
                            replaced_count += 1
                            print(f"      ✏️ {key} -> {value} en celda {cell.coordinate}")
                    
                    # Actualizar celda si hubo cambios
                    if modified_text != original_text:
                        cell.value = modified_text
    
    print(f"📝 Total reemplazos de texto: {replaced_count}")
    return replaced_count

def replace_placeholder_images_in_excel(workbook, placeholder_replacements: dict):
    """
    Sistema de placeholders para Excel - busca imágenes existentes y las reemplaza.
    Esto es más limitado que Word porque Excel no tiene la misma flexibilidad.
    """
    print("🖼️ Procesando reemplazos de placeholders en Excel...")
    
    if not placeholder_replacements:
        print("⚠️ No hay placeholders de imagen configurados para Excel")
        return 0
    
    # Verificar archivos de reemplazo
    valid_replacements = {}
    for name, info in placeholder_replacements.items():
        path = get_replacement_path(info)
        if path and os.path.exists(path):
            print(f"✅ {name} -> {path}")
            valid_replacements[name] = info
        else:
            print(f"❌ {name} -> {path} (NO EXISTE)")
    
    if not valid_replacements:
        print("❌ No hay archivos de reemplazo válidos")
        return 0
    
    total_replaced = 0
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"   📄 Procesando hoja: {sheet_name}")
        
        # Procesar imágenes existentes en la hoja
        images_to_remove = []
        images_to_add = []
        
        # Excel maneja las imágenes de manera diferente - están en ws._images
        if hasattr(ws, '_images') and ws._images:
            print(f"      🖼️ Encontradas {len(ws._images)} imágenes existentes")
            
            for img in ws._images:
                # Para cada imagen existente, intentar reemplazarla
                if valid_replacements:
                    placeholder_name, replacement_info = list(valid_replacements.items())[0]
                    replacement_path = get_replacement_path(replacement_info)
                    
                    print(f"      🔄 Reemplazando imagen en {img.anchor}")
                    
                    # Preparar nueva imagen
                    new_image_info = {
                        'path': replacement_path,
                        'anchor': img.anchor,
                        'original_width': getattr(img, 'width', None),
                        'original_height': getattr(img, 'height', None),
                        'replacement_config': replacement_info
                    }
                    
                    images_to_remove.append(img)
                    images_to_add.append(new_image_info)
                    total_replaced += 1
        
        # Remover imágenes originales
        for img in images_to_remove:
            if img in ws._images:
                ws._images.remove(img)
        
        # Añadir imágenes nuevas
        for img_info in images_to_add:
            try:
                add_replacement_image_to_excel(ws, img_info)
                print(f"      ✅ Imagen reemplazada exitosamente")
            except Exception as e:
                print(f"      ❌ Error reemplazando imagen: {e}")
        
        # También buscar marcadores de texto que indiquen dónde poner imágenes
        text_marker_replacements = replace_text_markers_with_images(ws, valid_replacements)
        total_replaced += text_marker_replacements
    
    print(f"🖼️ Total placeholders reemplazados: {total_replaced}")
    return total_replaced

def replace_text_markers_with_images(ws, valid_replacements):
    debug_excel_content(ws)
    """
    Busca marcadores de texto como {{LOGO}} y los reemplaza con imágenes.
    """
    replaced_count = 0
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                original_text = cell.value
                
                # Buscar marcadores de imagen en el texto
                for placeholder_name, replacement_info in valid_replacements.items():
                    # Buscar tanto el nombre del placeholder como marcadores de texto
                    markers_to_check = [
                        f"{{{{{placeholder_name.replace('.png', '').replace('placeholder_', '').upper()}}}}}",
                        placeholder_name,
                        f"<<{placeholder_name.replace('.png', '').replace('placeholder_', '').upper()}>>",
                        f"{{LOGO}}", # Marcador genérico común
                        f"{{IMAGEN}}"
                    ]
                    
                    for marker in markers_to_check:
                        if marker in original_text:
                            print(f"      🎯 Marcador encontrado: {marker} en {cell.coordinate}")
                            
                            # Limpiar el texto
                            cell.value = original_text.replace(marker, "").strip()
                            if not cell.value:
                                cell.value = None
                            
                            # Añadir imagen en la celda
                            try:
                                replacement_path = get_replacement_path(replacement_info)
                                img_info = {
                                    'path': replacement_path,
                                    'anchor': cell.coordinate,
                                    'replacement_config': replacement_info
                                }
                                add_replacement_image_to_excel(ws, img_info)
                                replaced_count += 1
                                print(f"      ✅ Imagen añadida en {cell.coordinate}")
                            except Exception as e:
                                print(f"      ❌ Error añadiendo imagen en {cell.coordinate}: {e}")
                            
                            break  # Solo procesar un marcador por celda
    
    return replaced_count

def add_replacement_image_to_excel(ws, img_info):
    """Añade una imagen de reemplazo a Excel."""
    replacement_path = img_info['path']
    anchor = img_info['anchor']
    config = img_info.get('replacement_config', {})
    
    # Crear imagen
    img = Image(replacement_path)
    
    # Calcular dimensiones
    if isinstance(config, dict):
        if 'width_pixels' in config and 'height_pixels' in config:
            img.width = config['width_pixels']
            img.height = config['height_pixels']
        elif 'width_cm' in config and 'height_cm' in config:
            # Convertir cm a pixels (aproximado: 96 DPI)
            img.width = int(config['width_cm'] * 96 / 2.54)
            img.height = int(config['height_cm'] * 96 / 2.54)
        elif config.get('maintain_aspect', False):
            # Calcular dimensiones manteniendo aspect ratio
            original_width = img_info.get('original_width')
            original_height = img_info.get('original_height')
            
            if original_width and original_height:
                img.width = original_width
                img.height = original_height
            else:
                # Dimensiones por defecto si no hay originales
                img.width = 100
                img.height = 100
                
                # Intentar mantener aspect ratio basado en la imagen real
                try:
                    with PILImage.open(replacement_path) as pil_img:
                        pil_width, pil_height = pil_img.size
                        aspect_ratio = pil_width / pil_height
                        
                        if aspect_ratio > 1:  # Más ancha
                            img.width = 150
                            img.height = int(150 / aspect_ratio)
                        else:  # Más alta
                            img.height = 150
                            img.width = int(150 * aspect_ratio)
                except:
                    pass  # Usar dimensiones por defecto si falla
    
    # Añadir imagen a la hoja
    ws.add_image(img, anchor)

def get_replacement_path(replacement_info):
    """Obtiene la ruta de la imagen de reemplazo."""
    if isinstance(replacement_info, str):
        return replacement_info
    elif isinstance(replacement_info, dict):
        return replacement_info.get('path')
    return None

def process_excel_file(input_path: str, output_path: str, replacements: dict, 
                      image_replacements: dict = None, placeholder_replacements: dict = None):
    """
    Procesa un archivo Excel con reemplazos de texto y placeholders de imagen.
    
    Args:
        input_path: Ruta del archivo de entrada
        output_path: Ruta del archivo de salida  
        replacements: Diccionario con reemplazos de texto
        image_replacements: Reemplazos de imágenes por marcadores de texto (compatibilidad)
        placeholder_replacements: Sistema nuevo de placeholders
    """
    if image_replacements is None:
        image_replacements = {}
    if placeholder_replacements is None:
        placeholder_replacements = {}
    
    try:
        print(f"\n📊 ========== PROCESANDO EXCEL: {os.path.basename(input_path)} ==========")
        wb = load_workbook(input_path)
        
        print(f"📋 Información del archivo:")
        print(f"   📄 Hojas: {len(wb.sheetnames)} ({', '.join(wb.sheetnames)})")
        
        text_replaced = 0
        images_replaced = 0
        
        # FASE 1: Reemplazos de texto
        if replacements:
            print(f"\n📝 FASE 1: Procesando {len(replacements)} reemplazos de texto...")
            text_replaced = replace_text_in_excel(wb, replacements)
        
        # FASE 2: Reemplazos de imágenes por texto (sistema legacy)
        if image_replacements:
            print(f"\n🖼️ FASE 2A: Procesando {len(image_replacements)} reemplazos de imagen por marcador...")
            # Convertir image_replacements al formato de placeholder_replacements
            for key, value in image_replacements.items():
                if key not in placeholder_replacements:
                    placeholder_replacements[f"marker_{key}"] = value
        
        # FASE 3: Sistema de placeholders
        if placeholder_replacements:
            print(f"\n🖼️ FASE 2B: Procesando {len(placeholder_replacements)} placeholders de imagen...")
            for name, info in placeholder_replacements.items():
                path = get_replacement_path(info)
                print(f"   📸 {name} → {os.path.basename(path) if path else 'N/A'}")
            
            images_replaced = replace_placeholder_images_in_excel(wb, placeholder_replacements)
        
        # Guardar archivo
        print(f"\n💾 Guardando archivo Excel...")
        wb.save(output_path)
        
        print(f"✅ Archivo Excel procesado exitosamente:")
        print(f"   📝 Reemplazos de texto: {text_replaced}")
        print(f"   🖼️ Reemplazos de imágenes: {images_replaced}")
        print(f"   📁 Guardado en: {output_path}")
        print("=" * 60 + "\n")
        
    except Exception as e:
        print(f"❌ Error procesando archivo Excel: {e}")
        import traceback
        traceback.print_exc()
        raise

# Función de utilidad para crear placeholders en Excel
def create_excel_template_with_placeholders(output_path, placeholders_info):
    """
    Crea un template de Excel con marcadores para placeholders.
    
    placeholders_info: {
        "{{LOGO}}": {"row": 1, "col": 1},
        "{{FIRMA}}": {"row": 10, "col": 5}
    }
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    print("📊 Creando template de Excel con placeholders...")
    
    for placeholder, position in placeholders_info.items():
        row = position.get('row', 1)
        col = position.get('col', 1)
        
        # Escribir el marcador en la celda
        cell = ws.cell(row=row, column=col)
        cell.value = placeholder
        
        # Opcional: dar formato especial a la celda
        from openpyxl.styles import Font, PatternFill
        cell.font = Font(bold=True, color="FF0000")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        print(f"   📍 {placeholder} en fila {row}, columna {col}")
    
    wb.save(output_path)
    print(f"✅ Template creado: {output_path}")

# Ejemplo de uso
if __name__ == "__main__":
    # Crear template de ejemplo
    create_excel_template_with_placeholders(
        "template_excel.xlsx",
        {
            "{{LOGO}}": {"row": 1, "col": 1},
            "{{NOMBRE}}": {"row": 2, "col": 2},
            "{{FIRMA}}": {"row": 10, "col": 5}
        }
    )
    
    # Ejemplo de procesamiento
    replacements = {
        "{{NOMBRE}}": "Juan Pérez",
        "{{EMPRESA}}": "Mi Empresa S.A."
    }
    
    placeholder_replacements = {
        "placeholder_logo.png": {
            "path": "mi_logo.png",
            "width_pixels": 100,
            "height_pixels": 50,
            "maintain_aspect": True
        }
    }
    
    process_excel_file(
        "template_excel.xlsx",
        "output_excel.xlsx",
        replacements,
        {},
        placeholder_replacements
    )